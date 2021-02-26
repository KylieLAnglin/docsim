# %%
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import matplotlib.pyplot as plt
from docsim.library import start
from docsim.library import analyze
from docsim.library import process_text

from openpyxl import load_workbook


# %% import

training = pd.read_csv(start.RAW_FILEPATH + "/validation/" + "human_codes.csv")
training["id"] = training.id.astype(str)

results = pd.read_csv(start.CLEAN_FILEPATH + "script_sims.csv")
results["id"] = results.id.astype(str)


# %%
training = training[
    ["study", "id", "training", "fidelity", "quality", "content"]
].merge(
    results,
    how="left",
    left_on=["study", "id"],
    right_on=["study", "id"],
    indicator=True,
)

# training = training[training.training == 0]


# %%
file = start.TABLE_FILEPATH + "validation.xlsx"
wb = load_workbook(file)
ws = wb.active
row = 4
col = 2

# %%
mod = smf.ols(formula="script_sim1 ~ + fidelity", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.fidelity.corr(training.script_sim1), 2
)
col = col + 1

training.fidelity.corr(training.script_sim1)

# %% New
mod = smf.ols(formula="script_sim2 ~ + fidelity", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.fidelity.corr(training.script_sim2), 2
)
col = col + 1

training.fidelity.corr(training.script_sim2)


# %% BEST
# THIS IS THE BEST
mod = smf.ols(formula="script_sim3 ~ + fidelity", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.fidelity.corr(training.script_sim3), 2
)
col = col + 1


# %% BEST
# THIS IS THE BEST
mod = smf.ols(formula="script_sim4 ~ + fidelity", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.fidelity.corr(training.script_sim4), 2
)
col = col + 1

plt.plot(training.fidelity, training.script_sim4, "o", color="black", alpha=0.1)
plt.xlabel("Fidelity Rubric Scores")
plt.ylabel("Script Similarity")
plt.savefig(start.TABLE_FILEPATH + "Validation Scatter Plot")

# %%

mod = smf.ols(formula="script_sim5 ~ + fidelity", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.fidelity.corr(training.script_sim5), 2
)
col = col + 1
wb.save(file)


# %%


# script_median = training.script_sim3.median()
fidelity_low = np.nanpercentile(training.fidelity, 25)
print("25th percentile fidelity = ", fidelity_low)
fidelity_high = np.nanpercentile(training.fidelity, 75)
print("75th percentile fidelity = ", fidelity_high)

script_low = np.nanpercentile(training.script_sim3, 25)
print("25th percentile script = ", script_low)
script_high = np.nanpercentile(training.script_sim3, 75)
print("75th percentile script = ", script_high)


script_median = 0.25
fidelity_median = training.fidelity.median()
print(len(training))
print(
    len(
        training[
            (training.fidelity > fidelity_high) & (training.script_sim3 > script_high)
        ]
    ),
    "high-fidelity, high-similarity",
)
print(
    len(
        training[
            (training.fidelity > fidelity_high) & (training.script_sim3 < script_high)
        ]
    ),
    "high-fidelity, low-similarity",
)

print(
    len(
        training[
            (training.fidelity < fidelity_low) & (training.script_sim3 < script_low)
        ]
    ),
    "low-fidelity, low-similarity",
)

print(
    len(
        training[
            (training.fidelity < fidelity_low) & (training.script_sim3 > fidelity_low)
        ]
    ),
    "low-fidelity, high-similarity",
)
# %%
file = start.TABLE_FILEPATH + "validation_content.xlsx"
wb = load_workbook(file)
ws = wb.active
row = 4
col = 2

# %%
mod = smf.ols(formula="script_sim1 ~ + content", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.content.corr(training.script_sim1), 2
)
col = col + 1

training.fidelity.corr(training.script_sim1)

# %% New
mod = smf.ols(formula="script_sim2 ~ + content", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.content.corr(training.script_sim2), 2
)
col = col + 1

training.fidelity.corr(training.script_sim2)

# %%
mod = smf.ols(formula="script_sim3 ~ + content", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.content.corr(training.script_sim3), 2
)
col = col + 1


# %% BEST
# THIS IS THE BEST
mod = smf.ols(formula="script_sim4 ~ + content", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.content.corr(training.script_sim4), 2
)
col = col + 1

plt.plot(training.fidelity, training.script_sim4, "o", color="black", alpha=0.1)

# %%

mod = smf.ols(formula="script_sim5 ~ + content", data=training)
res = mod.fit()
print(res.summary())
ws.cell(row=row, column=col).value = round(
    training.content.corr(training.script_sim5), 2
)
col = col + 1


# %%
wb.save(file)
