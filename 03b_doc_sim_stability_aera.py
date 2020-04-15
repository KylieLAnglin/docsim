#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from scipy import spatial
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.decomposition import TruncatedSVD
from sklearn.preprocessing import Normalizer


# In[6]:


clean_filepath = "/Users/kylieleblancKylie/domino/docsim/data/clean/aera/"
table_filepath = "/Users/kylieleblancKylie/domino/docsim/results/"


# In[7]:


docs = pd.read_csv(clean_filepath + 'text_aera_open.csv')
docs = docs.set_index('doc')
docs.sample(5)


# In[17]:


def similarity(df, matrix):
    for maindoc in list(matrix.index):
        pairwise_sim = []
        for doc in list(matrix.index):
            sim = 1 - spatial.distance.cosine(matrix.loc[maindoc], matrix.loc[doc])
            pairwise_sim.append(sim)
        average = (sum(pairwise_sim) - 1)/(len(pairwise_sim) - 1) # don't include relationship with self
        df.at[maindoc, 'doc_sim'] = average
    df.doc_sim.mean()
    return df


# In[18]:


file = table_filepath + 'AERA Results.xlsx'
wb = load_workbook(file)
ws = wb.active


# # No pre-processing

# In[19]:


results_corpus = docs.copy()
results_corpus.sample()


# In[20]:


matrix_corpus = pd.read_csv(clean_filepath + 'matrix_open.csv')
matrix_corpus = matrix_corpus.set_index('doc')
matrix_corpus.sample(5)


# In[21]:


#within_study_mean(results_corpus, 'spring', '2017-18', matrix_corpus)
#within_study_mean(results_corpus, 'fall', '2018-19', matrix_corpus)
similarity(results_corpus,  matrix_corpus)
value = results_corpus.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 2).value = value
wb.save(file)
results_corpus.sample(5)


# # Remove stop words

# In[22]:


results_corpus_stop = docs.copy()
results_corpus_stop.sample()


# In[23]:


matrix_corpus_stop = pd.read_csv(clean_filepath + 'matrix_open_stop.csv')
matrix_corpus_stop = matrix_corpus_stop.set_index('doc')
matrix_corpus_stop.sample(5)


# In[24]:


similarity(results_corpus_stop,matrix_corpus_stop)
value = results_corpus_stop.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 3).value = value
wb.save(file)
results_corpus_stop.sample(5)


# # Stop and Stem

# In[25]:


results_corpus_stop_stem = docs.copy()
results_corpus_stop_stem.sample()

matrix_corpus_stop_stem = pd.read_csv(clean_filepath + 'matrix_open_stop_stem.csv')
matrix_corpus_stop_stem = matrix_corpus_stop_stem.set_index('doc')
matrix_corpus_stop_stem.sample(5)

similarity(results_corpus_stop_stem, matrix_corpus_stop_stem)
value = results_corpus_stop_stem.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 4).value = value
wb.save(file)
results_corpus_stop_stem.sample(5)


# # Stop Stem Weight

# In[27]:


results_corpus_stop_stem_wgt = docs.copy()
results_corpus_stop_stem_wgt.sample()

matrix_corpus_stop_stem_wgt = pd.read_csv(clean_filepath + 'matrix_open_stop_stem_wgt.csv')
matrix_corpus_stop_stem_wgt = matrix_corpus_stop_stem_wgt.set_index('doc')
matrix_corpus_stop_stem_wgt.sample(5)

similarity(results_corpus_stop_stem_wgt, matrix_corpus_stop_stem_wgt)
value = results_corpus_stop_stem_wgt.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 5).value = value
wb.save(file)
results_corpus_stop_stem_wgt.sample(5)


# # LSA

# In[28]:


results_corpus_lsa = docs.copy()
results_corpus_lsa.sample()

matrix_corpus = pd.read_csv(clean_filepath + 'matrix_open.csv')
matrix_corpus = matrix_corpus.set_index('doc')
matrix_corpus.sample(5)


# In[29]:


lsa = TruncatedSVD(n_components = 100, random_state = 100)
lsa_fit = lsa.fit_transform(matrix_corpus)
lsa_fit = Normalizer(copy=False).fit_transform(lsa_fit)
print(lsa_fit.shape)
lsa_fit


# In[30]:


## Each LSA component is a linear combo of words
word_weights = pd.DataFrame(lsa.components_, columns = matrix_corpus.columns)
word_weights.head()


# In[31]:


## Each document is a linear combination of components
matrix_corpus_lsa = pd.DataFrame(lsa_fit, index = matrix_corpus.index, columns = word_weights.index)
matrix_corpus_lsa.sample(5)


# In[33]:


similarity(results_corpus_lsa,  matrix_corpus_lsa)
value = results_corpus_lsa.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 6).value = value
wb.save(file)
results_corpus_lsa.sample(5)


# # LSA and Weighting

# In[34]:


results_corpus_wgt_lsa = docs.copy()
results_corpus_wgt_lsa.sample()

matrix_corpus_wgt = pd.read_csv(clean_filepath + 'matrix_open_wgt.csv')
matrix_corpus_wgt = matrix_corpus_wgt.set_index('doc')
matrix_corpus_wgt.sample(5)

lsa = TruncatedSVD(n_components = 100, random_state = 100)
lsa_fit = lsa.fit_transform(matrix_corpus_wgt)
lsa_fit = Normalizer(copy=False).fit_transform(lsa_fit)
print(lsa_fit.shape)
lsa_fit

## Each LSA component is a linear combo of words
word_weights = pd.DataFrame(lsa.components_, columns = matrix_corpus_wgt.columns)
word_weights.head()

## Each document is a linear combination of components
matrix_corpus_wgt_lsa = pd.DataFrame(lsa_fit, index = matrix_corpus_wgt.index, columns = word_weights.index)
matrix_corpus_wgt_lsa.sample(5)

similarity(results_corpus_wgt_lsa, matrix_corpus_wgt_lsa)
value = results_corpus_wgt_lsa.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 7).value = value
wb.save(file)


# # Stop, Weight, and LSA

# In[35]:


results_corpus_stop_wgt_lsa = docs.copy()
results_corpus_stop_wgt_lsa.sample()

matrix_corpus_stop_wgt = pd.read_csv(clean_filepath + 'matrix_open_stop_wgt.csv')
matrix_corpus_stop_wgt = matrix_corpus_stop_wgt.set_index('doc')
matrix_corpus_stop_wgt.sample(5)

lsa = TruncatedSVD(n_components = 100, random_state = 100)
lsa_fit = lsa.fit_transform(matrix_corpus_stop_wgt)
lsa_fit = Normalizer(copy=False).fit_transform(lsa_fit)
print(lsa_fit.shape)
lsa_fit

## Each LSA component is a linear combo of words
word_weights = pd.DataFrame(lsa.components_, columns = matrix_corpus_stop_wgt.columns)
word_weights.head()

## Each document is a linear combination of components
matrix_corpus_stop_wgt_lsa = pd.DataFrame(lsa_fit, index = matrix_corpus_stop_wgt.index, columns = word_weights.index)
matrix_corpus_stop_wgt_lsa.sample(5)

similarity(results_corpus_stop_wgt_lsa, matrix_corpus_stop_wgt_lsa)
value = results_corpus_stop_wgt_lsa.doc_sim.mean().round(2)
print(value)
ws.cell(row= 2, column= 8).value = value
wb.save(file)


# In[ ]:




