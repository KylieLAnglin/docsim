# %%
import pandas as pd
import nltk
import openpyxl

from docsim.library import process_text


gi = pd.read_csv("/Users/kylie/generalinquirer.xls")
# %%
PATH = "/Users/kylie/Dropbox/Active/docsim/data/excerpts"

files = []
texts = []
for filename in os.listdir(PATH):
    if ".txt" in filename:
        text = open(PATH + "/" + filename, "r").read()
        files.append(filename)
        texts.append(text)

df = pd.DataFrame(list(zip(files, texts)), columns=["Filename", "text"])


df["text_lemmatized"] = [
    process_text.process_text_nltk(text, lower_case=True, remove_punct=True, lemma=True)
    for text in df.text
]

df["lemmas"] = [
    process_text.process_text_nltk(
        text, lower_case=True, remove_punct=True, lemma=True, string_or_list="list"
    )
    for text in df.text
]


# %% import TAACO results
taaco = pd.read_csv(
    "/Users/kylie/Dropbox/Active/docsim/data/excerpts/excerpt_results.csv"
)
taaco = taaco[
    [
        "Filename",
        "lemma_ttr",
        "adjacent_overlap_all_sent",
        "word2vec_1_all_sent",
        "coordinating_conjuncts",
        "all_causal",
    ]
]

df = df.merge(taaco, on=["Filename"], how="inner")
# %%
def determine_tense_input(sentence: str, tense_to_return: str):
    text = nltk.word_tokenize(sentence)
    tagged = nltk.pos_tag(text)

    tense = {}
    tense["future"] = len([word for word in tagged if word[1] == "MD"])
    tense["present"] = len(
        [word for word in tagged if word[1] in ["VBP", "VBZ", "VBG"]]
    )
    tense["past"] = len([word for word in tagged if word[1] in ["VBD", "VBN"]])
    verb_tense = tense[tense_to_return]
    return verb_tense


df["past"] = [determine_tense_input(text, "past") for text in df.text]
df["present"] = [determine_tense_input(text, "present") for text in df.text]
df["future"] = [determine_tense_input(text, "future") for text in df.text]


# %% General inquirer
def count_words_in_dict(list_of_words: list, dictionary_list: list):
    return len([i for i in list_of_words if i in dictionary_list])


def gi_search(gi_cat: str, df: pd.DataFrame):
    temp_df = gi.copy()
    temp_df = temp_df.rename(columns={gi_cat: "gi_cat"})
    cat_word_list = list(temp_df[temp_df.gi_cat == gi_cat].word.unique())

    cat_words = [set([i for i in lemmas if i in cat_word_list]) for lemmas in df.lemmas]

    cat_word_count = [
        count_words_in_dict(lemmas, cat_word_list) for lemmas in df.lemmas
    ]

    return (cat_words, cat_word_count)


df["strong_words"] = gi_search("Strong", df)[0]
df["strong_word_count"] = gi_search("Strong", df)[1]


df["weak_words"] = gi_search("Weak", df)[0]
df["weak_word_count"] = gi_search("Weak", df)[1]

df["active_words"] = gi_search("Active", df)[0]
df["active_word_count"] = gi_search("Active", df)[1]

df["passive_words"] = gi_search("Passive", df)[0]
df["passive_word_count"] = gi_search("Passive", df)[1]

df["goal_words"] = gi_search("Goal", df)[0]
df["goal_word_count"] = gi_search("Goal", df)[1]

# %%
df = df.set_index("Filename")
df.to_csv("/Users/kylie/Dropbox/Active/docsim/excerpt_description.csv")

df["strong_to_weak"] = df.strong_word_count / (
    df.strong_word_count + df.weak_word_count
)
df["active_to_passive"] = df.active_word_count / (
    df.active_word_count + df.passive_word_count
)
df["future_to_past"] = df.future / (df.future + df.past)
# %% Create formatted table
TABLE_PATH = "/Users/kylie/Dropbox/Active/docsim/data/excerpts/"
transcript = "01_1920_05_085_22c_Transcript"
file = TABLE_PATH + transcript + ".xlsx"
wb = openpyxl.Workbook()
ws = wb.active
transcript = transcript + ".txt"

ws.cell(row=1, column=1).value = "Example Output"
ws.cell(row=2, column=2).value = "Benchmark"
ws.cell(row=2, column=3).value = "Transcript Distance"
ws.cell(row=2, column=4).value = "Transcript Similarity"


def paste_values(row: int, label: str, value: str):
    ws.cell(row=row, column=1).value = label
    benchmark_value = df.loc["behavior1.txt", value]
    transcript_value = df.loc["01_1920_05_085_22c_Transcript.txt", value]

    ws.cell(row=row, column=2).value = round(benchmark_value, 3)
    diff = benchmark_value - transcript_value
    ws.cell(row=row, column=3).value = round(diff, 3)
    sim = 1 - abs(diff)
    ws.cell(row=row, column=4).value = round(sim, 3)

    wb.save(file)


paste_values(row=3, label="Type-Token Ratio", value="lemma_ttr")
paste_values(row=4, label="Average Adjacent Overlap", value="adjacent_overlap_all_sent")
paste_values(row=5, label="Average Word2Vec Value", value="word2vec_1_all_sent")
paste_values(row=6, label="Proportion causal Words", value="all_causal")
paste_values(row=7, label="Proportion strong vs weak words", value="strong_to_weak")
paste_values(
    row=8, label="Proportion active vs passive words", value="active_to_passive"
)
paste_values(row=9, label="Proportion future to past verbs", value="future_to_past")
# %%
TABLE_PATH = "/Users/kylie/Dropbox/Active/docsim/data/excerpts/"
transcript = "01_1920_05_021_22c_Transcript"
file = TABLE_PATH + transcript + ".xlsx"
wb = openpyxl.Workbook()
ws = wb.active
transcript = transcript + ".txt"

ws.cell(row=1, column=1).value = "Example Output"
ws.cell(row=2, column=2).value = "Benchmark"
ws.cell(row=2, column=3).value = "Transcript Distance"
ws.cell(row=2, column=4).value = "Transcript Similarity"


def paste_values(row: int, label: str, value: str):
    ws.cell(row=row, column=1).value = label
    benchmark_value = df.loc["behavior1.txt", value]
    transcript_value = df.loc["01_1920_05_021_22c_Transcript.txt", value]

    ws.cell(row=row, column=2).value = round(benchmark_value, 5)
    diff = benchmark_value - transcript_value
    ws.cell(row=row, column=3).value = round(diff, 5)
    sim = 1 - abs(diff)
    ws.cell(row=row, column=4).value = round(sim, 5)

    wb.save(file)


paste_values(row=3, label="Type-Token Ratio", value="lemma_ttr")
paste_values(row=4, label="Average Adjacent Overlap", value="adjacent_overlap_all_sent")
paste_values(row=5, label="Average Word2Vec Value", value="word2vec_1_all_sent")
paste_values(row=6, label="Proportion causal Words", value="all_causal")
paste_values(row=7, label="Proportion strong vs weak words", value="strong_to_weak")
paste_values(
    row=8, label="Proportion active vs passive words", value="active_to_passive"
)
paste_values(row=9, label="Proportion future to past verbs", value="future_to_past")
# %%
