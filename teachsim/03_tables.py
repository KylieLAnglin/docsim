# %%
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from docsim.library import start

# %%

results = pd.read_csv(start.CLEAN_FILEPATH + "script_sims.csv").set_index("study", "id")

# %%
file = start.TABLE_FILEPATH + "table1_fidelity.xlsx"
wb = load_workbook(file)
ws = wb.active

columns = [2, 3, 4, 5, 6, 7]

script_sims = [
    "script_sim1",
    "script_sim2",
    "script_sim3",
    "script_sim4",
    "script_sim5",
]


def script_sim_table(study: str, row: int):

    for script_sim, col in zip(script_sims, columns):
        value = round(
            results.loc[study][script_sim].mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value
    wb.save(file)


script_sim_table("spring2018", 3)
script_sim_table("spring2019", 4)
script_sim_table("fall2019TAP", 5)
script_sim_table("fall2017", 6)
script_sim_table("fall2018", 7)

results.loc["spring2018"]["script_sim3"].std()
results.loc["spring2019"]["script_sim3"].std()
results.loc["fall2019TAP"]["script_sim3"].std()
results.loc["fall2017"]["script_sim3"].std()
results.loc["fall2018"]["script_sim3"].std()

# %% Script Similarity


# %% Study Similarity - Spring 2019 Baseline

results = pd.read_csv(start.CLEAN_FILEPATH + "study_sims.csv").set_index("study", "id")

file = start.TABLE_FILEPATH + "table2_replicability.xlsx"
wb = load_workbook(file)
ws = wb.active

study_sims = [
    "rep1_study1",
    "rep2_study1",
    "rep3_study1",
    "rep4_study1",
    "rep5_study1",
]


def study_sim_table(study: str, row: int):

    for study_sim, col in zip(study_sims, columns):
        value = round(
            results.loc[study][study_sim].mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value
    wb.save(file)


study_sim_table("spring2018", 3)
study_sim_table("spring2019", 4)
study_sim_table("fall2019TAP", 5)
study_sim_table("fall2017", 6)
study_sim_table("fall2018", 7)


# %% Study Matrix

file = start.TABLE_FILEPATH + "table3_replicability_matrix.xlsx"
wb = load_workbook(file)
ws = wb.active


def study_matrix(study: str, row: int):
    study_sims = [
        "rep4_study1",
        "rep4_study2",
        "rep4_study3",
        "rep4_study4",
        "rep4_study5",
    ]

    cols = [2, 3, 4, 5, 6]

    for comp, col in zip(study_sims, cols):
        value = round(
            results.loc[study][comp].mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value
    wb.save(file)


study_matrix("spring2018", 3)
study_matrix("spring2019", 4)
study_matrix("fall2019TAP", 5)
study_matrix("fall2017", 6)
study_matrix("fall2018", 7)

# %%
