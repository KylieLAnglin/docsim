#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from library import start

# In[2]:

clean_filepath = start.clean_filepath
table_filepath = start.table_filepath

# In[3]:


docs = pd.read_csv(clean_filepath + 'text_transcripts.csv')
docs = docs.set_index('doc')
docs.sample(5)


# In[4]:


results = pd.read_csv(clean_filepath + 'results.csv')
results.sample(3)


# In[5]:


techniques = ['', '_stop', '_stop_wgt', '_stem_stop_wgt', '_lsa_wgt_stop']
columns = [2, 3, 4, 5, 6]


# In[6]:

# Script Similarity

file = table_filepath + 'table1_fidelity.xlsx'
wb = load_workbook(file)
ws = wb.active

row = 3
# Behavior
# Spring 2018
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2017-18') &
                           (results.semester == 'spring'))].script_sim.mean(),
                  2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1
print(len(results[((results.year == '2017-18') &
                           (results.semester == 'spring'))]))

# Spring 2019
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2018-19') &
                           (results.semester == 'spring'))].script_sim.mean(),
                  2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1
print(len(results[((results.year == '2018-19') &
                           (results.semester == 'spring'))]))

# Fall 2019
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2019-20') &
                           (results.semester == 'fall'))].script_sim.mean(),
                  2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1
print(len(results[((results.year == '2019-20') &
                           (results.semester == 'fall'))]))

# Feedback
# Fall 2017
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2017-18') &
                           (results.semester == 'fall'))].script_sim.mean(),
                  2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1
print(len(results[((results.year == '2017-18') &
                           (results.semester == 'fall'))]))

# Fall 2018
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2018-19') &
                           (results.semester == 'fall'))].script_sim.mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
print(len(results[((results.year == '2018-19') &
                           (results.semester == 'fall'))]))



# %% Study Similarity - Spring 2019 Baseline

file = table_filepath + 'table2_replicability.xlsx'
wb = load_workbook(file)
ws = wb.active

row = 3
# spring 2017
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2017-18') &
                           (results.semester == 'spring'))].study_sim_spring2017_18.mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1

# spring 2018
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2018-19') &
                           (results.semester == 'spring'))].study_sim_spring2017_18.mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1

# Spring 2018
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2019-20') &
                           (results.semester == 'fall'))].study_sim_spring2017_18.mean(), 2)
    ws.cell(row=5, column=col).value = value
wb.save(file)
row = row + 1

# Fall 2017
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2017-18') &
                           (results.semester == 'fall'))].study_sim_spring2017_18.mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row + 1

# Fall 2018
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    value = round(results[((results.year == '2018-19') &
                           (results.semester == 'fall'))].study_sim_spring2017_18.mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)


# %% Study Matrix
file = table_filepath + 'table3_replicability_matrix.xlsx'
wb = load_workbook(file)
ws = wb.active


results = pd.read_csv(clean_filepath + 'results_lsa_wgt_stop.csv')

comps = ['study_sim_spring2017_18', 'study_sim_spring2018_19',
         'study_sim_fall2019_20',
         'study_sim_fall2017_18', 'study_sim_fall2018_19']
cols = [2, 3, 4, 5, 6]

row = 3
for comp, col in zip(comps, cols):
    value = round(results[((results.year == '2017-18') &
                           (results.semester == 'spring'))][comp].mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row+1

for comp, col in zip(comps, cols):
    value = round(results[((results.year == '2018-19') &
                           (results.semester == 'spring'))][comp].mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row+1

for comp, col in zip(comps, cols):
    value = round(results[((results.year == '2019-20') &
                           (results.semester == 'fall'))][comp].mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row+1

for comp, col in zip(comps, cols):
    value = round(results[((results.year == '2017-18') &
                           (results.semester == 'fall'))][comp].mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row+1

for comp, col in zip(comps, cols):
    value = round(results[((results.year == '2018-19') &
                           (results.semester == 'fall'))][comp].mean(), 2)
    ws.cell(row=row, column=col).value = value
wb.save(file)
row = row+1


