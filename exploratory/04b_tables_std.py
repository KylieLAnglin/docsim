#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from library import start

# In[2]:

clean_filepath = start.CLEAN_FILEPATH
table_filepath = start.TABLE_FILEPATH

# In[3]:


docs = pd.read_csv(clean_filepath + 'text_transcripts.csv')
docs = docs.set_index('doc')
docs.sample(5)


# In[4]:


results = pd.read_csv(clean_filepath + 'results.csv')
results.sample(3)


# In[5]:


techniques = ['', '_stop', '_stop_wgt', '_stem_stop_wgt', '_lsa_wgt_stop']
columns = [2, 3, 4, 5, 6]


# In[6]:

# Script Similarity

file = table_filepath + 'table1b_fidelity_std.xlsx'
wb = load_workbook(file)
ws = wb.active

# Behavior
# Spring 2018
for tech, col in zip(techniques, columns):
    results = pd.read_csv(clean_filepath + 'results' + tech + '.csv')
    std = results.script_sim.std()
    print(round(std, 2))
    spring2018 = results[((results.year == '2017-18') &
                          (results.semester == 'spring'))]
    spring2019 = results[((results.year == '2018-19') &
                          (results.semester == 'spring'))]
    fall2019 = results[((results.year == '2018-19') &
                        (results.semester == 'fall'))]
    fall2017 = results[((results.year == '2017-18') &
                        (results.semester == 'fall'))]
    fall2018 = results[((results.year == '2018-19') &
                        (results.semester == 'fall'))]
    row = 3
    for study in [spring2018, spring2019, fall2019,
                  fall2017, fall2018]:
        value = round(study.script_sim.mean()/std, 2)
        ws.cell(row=row, column=col).value = value
        row = row + 1

wb.save(file)


# %%
