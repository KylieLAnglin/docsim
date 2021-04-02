# %%
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from docsim.library import start

# %%

script_sims = pd.read_csv(start.CLEAN_FILEPATH + "script_sims.csv").set_index(
    "study", "id"
)
study_sims = pd.read_csv(start.CLEAN_FILEPATH + "study_sims.csv").set_index(
    "study", "id"
)

# %% Functions


def script_sim_table(study: str, start_row: int):
    specifications = [
        "script_sim1",
        "script_sim2",
        "script_sim3",
        "script_sim4",
        "script_sim5",
    ]

    for specification, col in zip(specifications, columns):
        row = start_row
        value = round(
            script_sims.loc[study][specification].mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value

        row = row + 1

        value = round(
            script_sims.loc[study][specification].std(),
            2,
        )
        value = "[" + str(value) + "]"
        ws.cell(row=row, column=col).value = value

        row = row + 1

        low = round(
            script_sims.loc[study][specification].min(),
            2,
        )

        high = round(
            script_sims.loc[study][specification].max(),
            2,
        )

        value = "(" + str(low) + ", " + str(high) + ")"
        ws.cell(row=row, column=col).value = value

    wb.save(file)


def study_sim_table(study: str, study_suffix: str, start_row: int):
    specifications = ["rep" + str(i) + "_" + study_suffix for i in list(range(1, 6))]

    for specification, col in zip(specifications, columns):
        row = start_row
        value = round(
            study_sims.loc[study][specification].mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value

        row = row + 1

        value = round(
            study_sims.loc[study][specification].std(),
            2,
        )
        value = "[" + str(value) + "]"
        ws.cell(row=row, column=col).value = value

        row = row + 1

        low = round(
            study_sims.loc[study][specification].min(),
            2,
        )

        high = round(
            study_sims.loc[study][specification].max(),
            2,
        )

        value = "(" + str(low) + ", " + str(high) + ")"
        ws.cell(row=row, column=col).value = value

    wb.save(file)


columns = [2, 3, 4, 5, 6]


# %%

file = start.TABLE_FILEPATH + "Methodological Appendix Behavior Study 1.xlsx"
wb = load_workbook(file)
ws = wb.active

script_sim_table("spring2018", 3)
study_sim_table(study="spring2018", start_row=8, study_suffix="study1")

# %%
file = start.TABLE_FILEPATH + "Methodological Appendix Behavior Study 2.xlsx"
wb = load_workbook(file)
ws = wb.active

script_sim_table("spring2019", 3)
study_sim_table(study="spring2019", start_row=8, study_suffix="study2")

# %%

file = start.TABLE_FILEPATH + "Methodological Appendix Behavior Study 3.xlsx"
wb = load_workbook(file)
ws = wb.active

script_sim_table("fall2019TAP", 3)
study_sim_table(study="fall2019TAP", start_row=8, study_suffix="study3")

# %%
file = start.TABLE_FILEPATH + "Methodological Appendix Feedback Study 1.xlsx"
wb = load_workbook(file)
ws = wb.active

script_sim_table("fall2017", 3)
study_sim_table(study="fall2017", start_row=8, study_suffix="study4")

# %%
file = start.TABLE_FILEPATH + "Methodological Appendix Feedback Study 2.xlsx"
wb = load_workbook(file)
ws = wb.active

script_sim_table("fall2018", 3)
study_sim_table(study="fall2018", start_row=8, study_suffix="study5")

# %%


# %%
