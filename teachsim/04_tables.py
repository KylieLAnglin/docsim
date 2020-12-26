# %%
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from docsim.library import start

# %%

results = pd.read_csv(start.clean_filepath + "results.csv")

# %%

techniques = [
    "",
    "_stop",
    "_stop_wgt",
    "_stop_stem_wgt",
    "_stop_wgt_lsa",
    "_stop_stem_wgt_lsa",
]
columns = [2, 3, 4, 5, 6, 7]


# %% Script Similarity

file = start.table_filepath + "table1_fidelity.xlsx"
wb = load_workbook(file)
ws = wb.active


def script_sim_table(study: str, row: int):
    for tech, col in zip(techniques, columns):
        results = pd.read_csv(
            start.clean_filepath + "results" + tech + ".csv"
        ).set_index("study", "id")
        value = round(
            results.loc[study].script_sim.mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value
    wb.save(file)


script_sim_table("spring2018", 3)
script_sim_table("spring2019", 4)
script_sim_table("fall2019TAP", 5)
script_sim_table("fall2017", 6)
script_sim_table("fall2018", 7)


# %% Study Similarity - Spring 2019 Baseline
def study_sim_table(study: str, row: int):
    for tech, col in zip(techniques, columns):
        results = pd.read_csv(
            start.clean_filepath + "results" + tech + ".csv"
        ).set_index("study", "id")
        value = round(
            results.loc[study].sim_spring2018.mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value
    wb.save(file)


file = start.table_filepath + "table2_replicability.xlsx"
wb = load_workbook(file)
ws = wb.active

study_sim_table("spring2018", 3)
study_sim_table("spring2019", 4)
study_sim_table("fall2019TAP", 5)
study_sim_table("fall2017", 6)
study_sim_table("fall2018", 7)


# %% Study Matrix


def study_matrix(study: str, row: int):
    comps = [
        "sim_spring2018",
        "sim_spring2019",
        "sim_fall2019TAP",
        "sim_fall2017",
        "sim_fall2018",
    ]
    cols = [2, 3, 4, 5, 6]

    for comp, col in zip(comps, cols):
        value = round(
            results.loc[study][comp].mean(),
            2,
        )
        ws.cell(row=row, column=col).value = value
    wb.save(file)


file = start.table_filepath + "table3_replicability_matrix.xlsx"
wb = load_workbook(file)
ws = wb.active

results = pd.read_csv(start.clean_filepath + "results_stop_wgt_lsa.csv").set_index(
    "study", "id"
)

study_matrix("spring2018", 3)
study_matrix("spring2019", 4)
study_matrix("fall2019TAP", 5)
study_matrix("fall2017", 6)
study_matrix("fall2018", 7)
