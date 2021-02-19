# %%
# %%
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import matplotlib.pyplot as plt
from docsim.library import start
from docsim.library import analyze
from docsim.library import process_text

from openpyxl import load_workbook


# %% import


results = pd.read_csv(start.CLEAN_FILEPATH + "vectorizations.csv")
results["id"] = results["id"].astype(str)

transcripts = pd.read_csv(start.CLEAN_FILEPATH + "text_transcripts.csv")
transcripts["id"] = transcripts["id"].astype(str)


df = results.merge(
    transcripts[["study", "id", "clean_text"]],
    how="inner",
    left_on=["study", "id"],
    right_on=["study", "id"],
)

filename_col = 2
scenario_col = 3
similarity_col = 5
text_col = 6
# %%
# %%

file = start.TABLE_FILEPATH + "vignettes.xlsx"
wb = load_workbook(file)
ws = wb.active


# %%
def print_vignette_row(scenario: str, row: int, ascending: bool, location: int):
    ws.cell(row=row, column=scenario_col).value = scenario

    df_scenario = df[df.scenario == scenario]
    df_scenario = df_scenario.sort_values(by="script_sim4", ascending=ascending)

    filename = df_scenario.iloc[location]["filename"]
    ws.cell(row=row, column=filename_col).value = filename

    sim = df_scenario.iloc[location]["script_sim4"]
    ws.cell(row=row, column=similarity_col).value = round(sim, 2)

    text = df_scenario.iloc[location]["clean_text"].split()[0:200]
    text = " ".join(text)
    ws.cell(row=row, column=text_col).value = text
    wb.save(file)


row = 2

print_vignette_row(scenario="behavior", row=row, ascending=True, location=0)
row = row + 1

print_vignette_row(scenario="behavior", row=row, ascending=True, location=1)
row = row + 1

print_vignette_row(scenario="behavior", row=row, ascending=False, location=0)
row = row + 1

print_vignette_row(scenario="behavior", row=row, ascending=False, location=1)
row = row + 1

print_vignette_row(scenario="feedback", row=row, ascending=True, location=0)
row = row + 1

print_vignette_row(scenario="feedback", row=row, ascending=True, location=1)
row = row + 1

print_vignette_row(scenario="feedback", row=row, ascending=False, location=0)
row = row + 1

print_vignette_row(scenario="feedback", row=row, ascending=False, location=1)
row = row + 1

# %%
